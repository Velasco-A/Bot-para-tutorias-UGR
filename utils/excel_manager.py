import pandas as pd
import os
import sys
import logging
import traceback
from pathlib import Path
import sqlite3
from datetime import datetime
import openpyxl

# Añadir directorio raíz al path
sys.path.append(str(Path(__file__).parent.parent))
from db.queries import get_db_connection, get_o_crear_carrera

# Configurar logger
logger = logging.getLogger(__name__)

# Variables globales para almacenar datos
usuarios_excel = {}  # {email: {datos...}}
excel_cargado = False
excel_last_updated = None

def cargar_excel_en_memoria():
    """Carga todo el Excel en memoria una vez"""
    global usuarios_excel, excel_cargado, excel_last_updated
    
    try:
        # Buscar el Excel
        excel_path = Path(__file__).parent.parent / "data" / "usuarios.xlsx"
        
        if not os.path.exists(excel_path):
            print("❌ Excel no encontrado")
            return False
        
        # Cargar el Excel
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        # Obtener encabezados
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            if header:
                headers.append(header.strip())
            else:
                headers.append(f"Column_{col}")
        
        # Encontrar columna de email
        email_col = None
        for i, header in enumerate(headers):
            if 'mail' in header.lower():
                email_col = i + 1  # +1 porque las columnas en openpyxl empiezan en 1
                break
        
        if not email_col:
            print("❌ No se encontró columna de email")
            return False
        
        # Cargar todos los datos
        usuarios_excel.clear()
        total_usuarios = 0
        
        for row in range(2, sheet.max_row + 1):
            # Obtener email (columna encontrada)
            email_value = sheet.cell(row=row, column=email_col).value
            
            if not email_value:
                continue
                
            email = str(email_value).lower().strip()
            
            # Crear diccionario con todos los datos de la fila
            datos = {}
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if value:
                    datos[headers[col-1]] = str(value)
            
            # Guardar en el diccionario global
            usuarios_excel[email] = datos
            total_usuarios += 1
        
        print(f"✅ Excel cargado en memoria: {total_usuarios} usuarios")
        print(f"📧 Emails cargados: {list(usuarios_excel.keys())}")
        
        excel_cargado = True
        from datetime import datetime
        excel_last_updated = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        return True
        
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        print(traceback.format_exc())
        return False

def verificar_email_en_excel(email):
    """Verifica si un email está en los datos cargados (muy simple ahora)"""
    global usuarios_excel, excel_cargado
    
    # Si no está cargado, intentar cargar
    if not excel_cargado:
        cargar_excel_en_memoria()
    
    # Normalizar email
    email_norm = email.lower().strip()
    
    # Verificar si existe
    existe = email_norm in usuarios_excel
    print(f"🔍 Verificando '{email_norm}': {'✅ ENCONTRADO' if existe else '❌ NO ENCONTRADO'}")
    return existe

def obtener_datos_por_email(email):
    """Obtiene los datos de un usuario por su email"""
    global usuarios_excel, excel_cargado
    
    # Si no está cargado, intentar cargar
    if not excel_cargado:
        cargar_excel_en_memoria()
    
    # Normalizar email
    email_norm = email.lower().strip()
    
    # Retornar datos o None
    return usuarios_excel.get(email_norm)

def cargar_excel():
    """Carga datos del Excel a la base de datos"""
    try:
        # Buscar el Excel en la carpeta data y en raíz
        excel_path = None
        posibles_rutas = [
            Path(__file__).parent.parent / "data" / "usuarios.xlsx",
            Path(__file__).parent.parent / "usuarios.xlsx"
        ]
        
        for ruta in posibles_rutas:
            if ruta.exists():
                excel_path = ruta
                break
        
        if not excel_path:
            print(f"❌ Excel no encontrado en rutas: {[str(p) for p in posibles_rutas]}")
            return False
        
        print(f"📄 Cargando Excel desde: {excel_path}")
        
        # Cargar el Excel con todos los datos como strings para evitar conversiones automáticas
        df = pd.read_excel(excel_path, dtype=str)
        print(f"📊 Excel cargado con {len(df)} filas y columnas: {list(df.columns)}")
        
        # Mostrar primeras filas para diagnóstico
        print(f"Muestra de datos:\n{df.head(1).to_string()}")
        
        # Contadores para estadísticas
        usuarios_procesados = 0
        asignaturas_procesadas = 0
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Procesar cada fila del Excel
        for i, row in df.iterrows():
            try:
                # Verificar datos mínimos necesarios
                if 'Email' not in df.columns:
                    print(f"❌ Error: Columna 'Email' no encontrada en el Excel. Columnas disponibles: {list(df.columns)}")
                    return False
                    
                nombre = row.get('Nombre', '').strip()
                email = row.get('Email', '').strip().lower()
                
                if not nombre or not email:
                    print(f"⚠️ Fila {i+1}: Saltada por falta de nombre o email")
                    continue
                
                # Datos adicionales
                apellidos = row.get('Apellidos', '').strip()
                dni = row.get('DNI', '').strip()
                tipo = row.get('Tipo', 'estudiante').strip().lower()
                carrera = row.get('Carrera', '').strip()
                
                # Comprobar si el usuario ya existe
                cursor.execute("SELECT Id_usuario FROM Usuarios WHERE Email_UGR = ?", (email,))
                usuario_existente = cursor.fetchone()
                
                if usuario_existente:
                    # Actualizar usuario existente
                    cursor.execute("""
                        UPDATE Usuarios 
                        SET Nombre=?, Apellidos=?, DNI=?, Tipo=?, Carrera=?
                        WHERE Email_UGR=?
                    """, (nombre, apellidos, dni, tipo, carrera, email))
                    user_id = usuario_existente[0]
                    print(f"✓ Usuario actualizado: {nombre} ({email})")
                else:
                    # Crear nuevo usuario
                    cursor.execute("""
                        INSERT INTO Usuarios (Nombre, Apellidos, DNI, Email_UGR, Tipo, Carrera)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (nombre, apellidos, dni, email, tipo, carrera))
                    user_id = cursor.lastrowid
                    print(f"✓ Usuario creado: {nombre} ({email}) con ID: {user_id}")
                
                usuarios_procesados += 1
                
                # Procesar carrera
                if carrera:
                    carrera_id = get_o_crear_carrera(carrera)
                else:
                    carrera_id = None
                
                # Procesar asignaturas - buscando en ambas columnas posibles
                asignaturas = []
                for col_name in ['Asignaturas', 'Asignatura']:
                    if col_name in df.columns and not pd.isna(row.get(col_name)):
                        asig_text = str(row.get(col_name)).strip()
                        if ";" in asig_text:
                            asignaturas.extend([a.strip() for a in asig_text.split(";")])
                        elif "," in asig_text:
                            asignaturas.extend([a.strip() for a in asig_text.split(",")])
                        else:
                            asignaturas.append(asig_text)
                
                # Si hay columnas ST, SRC, RIM como booleanos, convertirlas a asignaturas
                for asig_col in ['ST', 'SRC', 'RIM']:
                    if asig_col in df.columns and str(row.get(asig_col)).lower() in ['1', 'true', 'yes', 'si', 'sí']:
                        asignaturas.append(asig_col)
                
                # Procesar cada asignatura
                for asig_nombre in asignaturas:
                    if not asig_nombre.strip():
                        continue
                    
                    # Buscar o crear asignatura
                    cursor.execute("SELECT Id_asignatura FROM Asignaturas WHERE Nombre = ?", (asig_nombre,))
                    asig = cursor.fetchone()
                    
                    if not asig:
                        cursor.execute("""
                            INSERT INTO Asignaturas (Nombre, Id_carrera) 
                            VALUES (?, ?)
                        """, (asig_nombre, carrera_id))
                        asig_id = cursor.lastrowid
                    else:
                        asig_id = asig[0]
                        
                        # Actualizar carrera si es necesario
                        if carrera_id:
                            cursor.execute("""
                                UPDATE Asignaturas SET Id_carrera = ? 
                                WHERE Id_asignatura = ? AND (Id_carrera IS NULL OR Id_carrera = '')
                            """, (carrera_id, asig_id))
                    
                    # Crear matrícula
                    cursor.execute("""
                        INSERT OR IGNORE INTO Matriculas (Id_usuario, Id_asignatura, Tipo) 
                        VALUES (?, ?, ?)
                    """, (user_id, asig_id, tipo))
                    
                    asignaturas_procesadas += 1
                    print(f"  ✓ Asignatura: {asig_nombre} - ID: {asig_id}")
            
            except Exception as e:
                print(f"❌ Error en fila {i+1}: {e}")
                continue
        
        conn.commit()
        conn.close()
        global excel_last_updated
        excel_last_updated = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        print(f"✅ Excel cargado: {usuarios_procesados} usuarios, {asignaturas_procesadas} asignaturas")
        return True
        
    except Exception as e:
        print(f"❌ Error al cargar Excel: {e}")
        print(traceback.format_exc())
        return False

def buscar_usuario_por_email(email):
    """Busca un usuario por su email en los datos del Excel"""
    from db.queries import get_db_connection
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    email = email.lower().strip()
    cursor.execute("SELECT * FROM Usuarios WHERE Email_UGR = ?", (email,))
    result = cursor.fetchone()
    
    conn.close()
    return result

def get_last_updated():
    """Retorna fecha de última actualización de datos"""
    return excel_last_updated

def importar_datos_por_email(email):
    """Importa los datos de un usuario desde el Excel por su email"""
    try:
        # Buscar el Excel en múltiples ubicaciones
        excel_path = None
        posibles_rutas = [
            Path(__file__).parent.parent / "data" / "usuarios.xlsx",
            Path(__file__).parent.parent / "usuarios.xlsx",
        ]
        
        for ruta in posibles_rutas:
            if ruta.exists():
                excel_path = ruta
                break
        
        if not excel_path:
            print(f"❌ Excel no encontrado para importación")
            return False
            
        # Cargar el Excel
        print(f"📄 Buscando email '{email}' en: {excel_path}")
        df = pd.read_excel(excel_path, dtype=str)
        
        # Verificar columnas
        if 'Email' not in df.columns:
            print(f"❌ Columna 'Email' no encontrada en Excel. Columnas: {list(df.columns)}")
            return False
        
        # Preparar email para búsqueda y mostrar todos los emails
        email_norm = email.lower().strip()
        df['Email'] = df['Email'].astype(str).str.lower().str.strip()
        
        # Diagnóstico detallado
        print(f"🔍 Buscando: '{email_norm}'")
        print(f"📧 Emails en Excel: {df['Email'].tolist()}")
        
        # Buscar usuario
        user_row = df[df['Email'] == email_norm]
        
        if user_row.empty:
            print(f"❌ Email '{email_norm}' no encontrado en el Excel")
            return False
            
        print(f"✅ Email encontrado, procesando datos...")
        
        # Extraer datos del usuario
        user_data = user_row.iloc[0]
        
        # Crear usuario en DB
        from db.queries import create_user, update_user
        
        # Crear nuevo usuario o actualizar existente
        user = buscar_usuario_por_email(email)
        
        if user:
            user_id = user['Id_usuario']
            update_user(
                user_id,
                Nombre=user_data.get('Nombre'),
                Apellidos=user_data.get('Apellidos'),
                Carrera=user_data.get('Carrera'),
                Tipo=user_data.get('Tipo', 'estudiante')
            )
            print(f"✓ Usuario actualizado: {user_id}")
        else:
            user_id = create_user(
                nombre=user_data.get('Nombre'),
                apellidos=user_data.get('Apellidos'),
                tipo=user_data.get('Tipo', 'estudiante'),
                email=email,
                telegram_id=None,  # Esto se actualizará después
                dni=user_data.get('DNI', '')
            )
            print(f"✓ Usuario creado: {user_id}")
        
        # Procesar carrera
        carrera = user_data.get('Carrera')
        if carrera:
            carrera_id = get_o_crear_carrera(carrera)
            
            # Procesar asignaturas
            asignaturas = []
            for col_name in ['Asignaturas', 'Asignatura']:
                if col_name in user_data.index and not pd.isna(user_data.get(col_name)):
                    asig_text = str(user_data.get(col_name)).strip()
                    if ";" in asig_text:
                        asignaturas.extend([a.strip() for a in asig_text.split(";")])
                    elif "," in asig_text:
                        asignaturas.extend([a.strip() for a in asig_text.split(",")])
                    else:
                        asignaturas.append(asig_text)
            
            # También procesar columnas ST, SRC, RIM si existen
            for col in ['ST', 'SRC', 'RIM']:
                if col in user_data and str(user_data.get(col)).lower() in ['1', 'true', 'yes', 'si', 'sí']:
                    asignaturas.append(col)
            
            # Añadir cada asignatura
            for asig_nombre in asignaturas:
                if not asig_nombre.strip():
                    continue
                    
                # Procesamiento especial para asignaturas separadas por comas en el mismo campo
                asig_parts = [a.strip() for a in asig_nombre.split(',') if a.strip()]
                for asig in asig_parts:
                    conn = get_db_connection()
                    cursor = conn.cursor()
                    
                    # Buscar o crear asignatura
                    cursor.execute("SELECT Id_asignatura FROM Asignaturas WHERE Nombre = ?", (asig,))
                    asig_row = cursor.fetchone()
                    
                    if not asig_row:
                        cursor.execute("""
                            INSERT INTO Asignaturas (Nombre, Id_carrera) 
                            VALUES (?, ?)
                        """, (asig, carrera_id))
                        asig_id = cursor.lastrowid
                    else:
                        asig_id = asig_row[0]
                    
                    # Crear matrícula
                    cursor.execute("""
                        INSERT OR IGNORE INTO Matriculas (Id_usuario, Id_asignatura, Tipo)
                        VALUES (?, ?, ?)
                    """, (user_id, asig_id, user_data.get('Tipo', 'estudiante')))
                    
                    conn.commit()
                    conn.close()
                    print(f"  ✓ Asignatura registrada: {asig}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error al importar datos: {e}")
        print(traceback.format_exc())
        return False

def verificar_excel_disponible():
    """Verifica si el archivo Excel está disponible sin cargarlo"""
    try:
        posibles_rutas = [
            Path(__file__).parent.parent / "data" / "usuarios.xlsx",
            Path(__file__).parent.parent / "usuarios.xlsx",
        ]
        
        for ruta in posibles_rutas:
            if ruta.exists():
                print(f"✅ Excel encontrado en: {ruta}")
                return True
                
        return False
        
    except Exception as e:
        print(f"❌ Error al verificar Excel: {e}")
        return False

# Para pruebas directas
if __name__ == "__main__":
    pass
